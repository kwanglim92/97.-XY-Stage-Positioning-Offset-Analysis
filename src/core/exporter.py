"""
exporter.py — CSV / Excel 내보내기
"""

import os
import csv
import collections
from typing import Optional

from core.statistics import (classify_row, spec_bounds, axis_reference_means,
                             active_anomaly_kinds, ANOMALY_FAILED,
                             ANOMALY_OUTLIER, ANOMALY_SPEC)


def export_combined_csv(data: list, output_path: str,
                        delimiter: str = '\t') -> str:
    """Analysis.txt 대체 — 배치 데이터를 CSV/TSV로 내보내기

    Args:
        data: batch_load() 결과
        output_path: 출력 파일 경로
        delimiter: 구분자 ('\\t' = TSV, ',' = CSV)

    Returns:
        저장된 파일 경로
    """
    if not data:
        return ''

    header = ['Foldername', 'Filename', 'Site ID', 'Site X', 'Site Y',
              'Point No', 'X (um)', 'Y (um)', 'Method ID', 'State',
              'Valid', 'HZ1_O (nm)', 'HZ1_O_Valid']

    with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f, delimiter=delimiter)
        writer.writerow(header)

        for r in data:
            writer.writerow([
                r.get('lot_name', ''),
                r.get('filename', ''),
                r.get('site_id', ''),
                r.get('site_x', ''),
                r.get('site_y', ''),
                r.get('point_no', ''),
                r.get('x_um', ''),
                r.get('y_um', ''),
                r.get('method', ''),
                r.get('state', ''),
                'TRUE' if r.get('valid', True) else 'FALSE',
                r.get('value', ''),
                'TRUE' if r.get('value_valid', True) else 'FALSE',
            ])

    return output_path


def export_statistics_csv(stats: list, output_path: str) -> str:
    """그룹별 통계를 CSV로 내보내기

    Args:
        stats: analyzer.compute_group_statistics() 결과
    """
    if not stats:
        return ''

    header = ['Group', 'Count', 'Mean', 'Stdev', 'Min', 'Max', 'Range']

    with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(header)
        for s in stats:
            writer.writerow([
                s.get('group', ''),
                s.get('count', 0),
                s.get('mean', 0),
                s.get('stdev', 0),
                s.get('min', 0),
                s.get('max', 0),
                s.get('range', 0),
            ])

    return output_path


# ──────────────────────────────────────────────
# 데이터 품질 리포트 — 어느 데이터가 어떻게 이상한지
# ──────────────────────────────────────────────

ANOMALY_HEADERS = ['Recipe', 'Round', 'Lot', 'Site ID', 'Die X', 'Die Y',
                   'Axis', 'Point No', 'State', 'Valid', 'HZ1_O (nm)',
                   '편차 (nm)', '기준평균 (nm)', '이상 유형', 'LSL', 'USL',
                   '원본 CSV', 'Lot 폴더']


def collect_anomalies(recipe_results: list, spec_limits: dict = None) -> list:
    """전체 Recipe에서 이상 측정 행을 추적정보와 함께 모은다.

    Returns:
        [{'recipe','round','lot','site_id','die_x','die_y','axis','point_no',
          'state','valid','value','kinds','lsl','usl','csv_path','lot_dir'}, ...]
    """
    rows = []
    for result in recipe_results or []:
        recipe = result.get('short_name') or result.get('recipe', '')
        round_name = result.get('round', '')
        round_path = result.get('round_path', '')
        # 화면 필터·PASS/FAIL과 같은 편차 기준을 쓴다 (Recipe별로 0점을 따로 잡는다)
        axis_means = axis_reference_means(result.get('raw_data', []))

        for r in result.get('raw_data', []):
            kinds = classify_row(r, spec_limits, recipe, axis_means=axis_means)
            if not kinds:
                continue

            lot = r.get('lot_name', '')
            lot_dir = os.path.join(round_path, lot) if round_path and lot else ''
            filename = r.get('filename', '')
            axis = (r.get('method') or '').upper()
            lsl, usl = spec_bounds(spec_limits, recipe, axis)
            ref = axis_means.get(axis)
            value = r.get('value', '')
            deviation = (round(value - ref, 3)
                         if ref is not None and isinstance(value, (int, float))
                         and value == value else '')

            rows.append({
                'deviation': deviation,
                'ref_mean': round(ref, 3) if ref is not None else '',
                'recipe': recipe,
                'round': round_name,
                'lot': lot,
                'site_id': r.get('site_id', ''),
                'die_x': r.get('site_x', ''),
                'die_y': r.get('site_y', ''),
                'axis': r.get('method', ''),
                'point_no': r.get('point_no', ''),
                'state': r.get('state', ''),
                'valid': 'TRUE' if r.get('valid', True) else 'FALSE',
                'value': r.get('value', ''),
                'kinds': ', '.join(kinds),
                'lsl': lsl if lsl is not None else '',
                'usl': usl if usl is not None else '',
                'csv_path': os.path.join(lot_dir, filename) if lot_dir and filename else filename,
                'lot_dir': lot_dir,
            })
    return rows


def build_quality_summary(recipe_results: list, anomalies: list) -> list:
    """Recipe별 요약 — 총 측정 대비 이상 건수."""
    by_recipe = collections.defaultdict(lambda: collections.Counter())
    for a in anomalies:
        for kind in a['kinds'].split(', '):
            by_recipe[a['recipe']][kind] += 1

    summary = []
    for result in recipe_results or []:
        recipe = result.get('short_name') or result.get('recipe', '')
        c = by_recipe.get(recipe, collections.Counter())
        total = len(result.get('raw_data', []))
        failed = c.get(ANOMALY_FAILED, 0)
        summary.append({
            'recipe': recipe,
            'round': result.get('round', ''),
            'total': total,
            'failed': failed,
            'outlier': c.get(ANOMALY_OUTLIER, 0),
            'spec': c.get(ANOMALY_SPEC, 0),
            'file_issues': len(result.get('data_warnings', [])),
            'failed_pct': round(failed / total * 100, 2) if total else 0,
            'load_error': result.get('error', ''),
        })
    return summary


def build_file_issues(recipe_results: list) -> list:
    """요약 CSV의 비정상 셀 경고 (예: 장비가 NaN을 U+FFFD로 기록)."""
    issues = []
    for result in recipe_results or []:
        recipe = result.get('short_name') or result.get('recipe', '')
        for msg in result.get('data_warnings', []):
            issues.append({'recipe': recipe, 'message': msg})
        for msg in result.get('load_errors', []):
            issues.append({'recipe': recipe, 'message': f'[로드 실패] {msg}'})
    return issues


def _group_counts(anomalies: list, key: str) -> list:
    """key(lot/site_id)별 유형 집계."""
    table = collections.defaultdict(lambda: collections.Counter())
    for a in anomalies:
        ident = (a['recipe'], a[key])
        for kind in a['kinds'].split(', '):
            table[ident][kind] += 1
            table[ident]['합계'] += 1

    rows = [{'recipe': rec, 'key': val,
             'failed': c.get(ANOMALY_FAILED, 0),
             'outlier': c.get(ANOMALY_OUTLIER, 0),
             'spec': c.get(ANOMALY_SPEC, 0),
             'total': c.get('합계', 0)}
            for (rec, val), c in table.items()]
    rows.sort(key=lambda r: (-r['total'], r['recipe'], str(r['key'])))
    return rows


def export_excel_report(data: list, stats: dict, trend: list,
                        output_path: str,
                        recipe_results: list = None,
                        spec_limits: dict = None,
                        log_rows: list = None) -> str:
    """Excel 리포트 (openpyxl 사용)

    시트 구성:
        1. Raw Data — 선택된 Step의 측정 데이터
        2. Statistics — 그룹별 통계
        3. Trend — Lot별 트렌드
        recipe_results/log_rows를 넘기면 아래 시트가 추가된다 (전체 Recipe 기준):
        4. Summary / Anomalies / By Lot / By Die / File Issues / System Log
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        # openpyxl 없으면 CSV로 대체
        return export_combined_csv(data, output_path.replace('.xlsx', '.csv'))

    wb = Workbook()

    # 공통 스타일
    header_font = Font(name='맑은 고딕', bold=True, size=10)
    header_fill = PatternFill(start_color='1565C0', end_color='1565C0',
                              fill_type='solid')
    header_font_white = Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF')
    data_font = Font(name='맑은 고딕', size=9)
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )

    # === Sheet 1: Raw Data ===
    ws1 = wb.active
    ws1.title = 'Raw Data'

    headers_raw = ['Lot', 'Filename', 'Site ID', 'Site X', 'Site Y',
                   'Point No', 'X (um)', 'Y (um)', 'Method', 'State',
                   'Valid', 'HZ1_O (nm)', 'Outlier']

    for col, h in enumerate(headers_raw, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for row_idx, r in enumerate(data, 2):
        values = [
            r.get('lot_name', ''), r.get('filename', ''),
            r.get('site_id', ''), r.get('site_x', ''),
            r.get('site_y', ''), r.get('point_no', ''),
            r.get('x_um', ''), r.get('y_um', ''),
            r.get('method', ''), r.get('state', ''),
            'TRUE' if r.get('valid', True) else 'FALSE',
            r.get('value', ''),
            'YES' if r.get('is_outlier', False) else '',
        ]
        for col, val in enumerate(values, 1):
            cell = ws1.cell(row=row_idx, column=col, value=val)
            cell.font = data_font
            cell.border = thin_border

    # 열 너비 조정
    for col in range(1, len(headers_raw) + 1):
        ws1.column_dimensions[chr(64 + col) if col <= 26 else 'A'].width = 14

    # === Sheet 2: Statistics ===
    if stats.get('overall'):
        ws2 = wb.create_sheet('Statistics')
        stat_headers = ['항목', '값']
        for col, h in enumerate(stat_headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill

        overall = stats['overall']
        stat_rows = [
            ('전체 데이터 수', overall.get('count', 0)),
            ('평균 (Mean)', overall.get('mean', 0)),
            ('표준편차 (Stdev)', overall.get('stdev', 0)),
            ('변동계수 (CV%)', overall.get('cv_percent', 0)),
        ]
        lot_var = stats.get('lot_variation', {})
        if lot_var:
            stat_rows.extend([
                ('', ''),
                ('--- Lot간 변동 ---', ''),
                ('Lot 수', lot_var.get('count', 0)),
                ('평균의 평균', lot_var.get('mean_of_means', 0)),
                ('평균의 표준편차', lot_var.get('stdev_of_means', 0)),
                ('평균의 범위', lot_var.get('range_of_means', 0)),
            ])

        for row_idx, (label, val) in enumerate(stat_rows, 2):
            ws2.cell(row=row_idx, column=1, value=label).font = data_font
            ws2.cell(row=row_idx, column=2, value=val).font = data_font

    # === Sheet 3: Trend ===
    if trend:
        ws3 = wb.create_sheet('Trend')
        trend_headers = ['Lot', 'Index', 'Count', 'Mean', 'Stdev', 'Min', 'Max']
        for col, h in enumerate(trend_headers, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill

        for row_idx, t in enumerate(trend, 2):
            values = [t['lot_name'], t['lot_index'], t['count'],
                      t['mean'], t['stdev'], t['min'], t['max']]
            for col, val in enumerate(values, 1):
                ws3.cell(row=row_idx, column=col, value=val).font = data_font

    # === Sheet 4~: 데이터 품질 / 로그 (전달된 경우에만) ===
    write_quality_sheets(wb, recipe_results, spec_limits, log_rows)

    wb.save(output_path)
    return output_path


# ──────────────────────────────────────────────
# 품질 리포트 시트 작성
# ──────────────────────────────────────────────

def _make_styles():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    return {
        'header_font': Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF'),
        'header_fill': PatternFill(start_color='1565C0', end_color='1565C0',
                                   fill_type='solid'),
        'data_font': Font(name='맑은 고딕', size=9),
        'bad_font': Font(name='맑은 고딕', size=9, color='C62828'),
        'align': Alignment(horizontal='center'),
        'border': Border(*[Side(style='thin', color='D0D0D0')] * 4),
    }


def _write_table(ws, headers: list, rows: list, st: dict,
                 width: int = 16, highlight=None):
    """헤더 + 행을 채우고 틀 고정·자동필터를 건다.

    highlight: row(dict/list)를 받아 True면 빨간 글씨로 표시하는 콜백.
    """
    from openpyxl.utils import get_column_letter

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = st['header_font']
        cell.fill = st['header_fill']
        cell.alignment = st['align']
        cell.border = st['border']

    for row_idx, values in enumerate(rows, 2):
        bad = bool(highlight and highlight(values))
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = st['bad_font'] if bad else st['data_font']
            cell.border = st['border']

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = 'A2'
    if rows:
        ws.auto_filter.ref = (f'A1:{get_column_letter(len(headers))}'
                              f'{len(rows) + 1}')


def write_quality_sheets(wb, recipe_results: list = None,
                         spec_limits: dict = None,
                         log_rows: list = None) -> None:
    """워크북에 데이터 품질·로그 시트를 덧붙인다 (인자가 없으면 아무것도 하지 않음)."""
    st = _make_styles()

    if recipe_results:
        anomalies = collect_anomalies(recipe_results, spec_limits)

        # Spec 초과 판정이 꺼져 있으면 관련 열도 내보내지 않는다 (화면과 동일 기준)
        show_spec = ANOMALY_SPEC in active_anomaly_kinds()

        # Summary
        ws = wb.create_sheet('Summary')
        headers = ['Recipe', 'Round', '총 측정', '측정 실패', '측정 실패 %', '이상치']
        if show_spec:
            headers.append('Spec 초과')
        headers += ['파일 이상', '로드 오류']
        _write_table(
            ws, headers,
            [[s['recipe'], s['round'], s['total'], s['failed'], s['failed_pct'],
              s['outlier']] + ([s['spec']] if show_spec else [])
             + [s['file_issues'], s['load_error']]
             for s in build_quality_summary(recipe_results, anomalies)],
            st, highlight=lambda v: bool(v[3]) or bool(v[-1]))

        # Anomalies — 상세 목록 + 추적정보
        ws = wb.create_sheet('Anomalies')
        _write_table(
            ws, ANOMALY_HEADERS,
            [[a['recipe'], a['round'], a['lot'], a['site_id'], a['die_x'],
              a['die_y'], a['axis'], a['point_no'], a['state'], a['valid'],
              a['value'], a['deviation'], a['ref_mean'], a['kinds'],
              a['lsl'], a['usl'], a['csv_path'], a['lot_dir']]
             for a in anomalies],
            st, width=18, highlight=lambda v: ANOMALY_FAILED in str(v[13]))

        # 집계
        for title, key in (('By Lot', 'lot'), ('By Die', 'site_id')):
            ws = wb.create_sheet(title)
            _write_table(
                ws,
                ['Recipe', 'Lot' if key == 'lot' else 'Site ID',
                 '측정 실패', '이상치'] + (['Spec 초과'] if show_spec else []) + ['합계'],
                [[g['recipe'], g['key'], g['failed'], g['outlier']]
                 + ([g['spec']] if show_spec else []) + [g['total']]
                 for g in _group_counts(anomalies, key)],
                st, width=18)

        # 파일 이상 (요약 CSV 비정상 셀)
        ws = wb.create_sheet('File Issues')
        _write_table(ws, ['Recipe', '내용'],
                     [[i['recipe'], i['message']]
                      for i in build_file_issues(recipe_results)],
                     st, width=60)

    if log_rows:
        ws = wb.create_sheet('System Log')
        _write_table(ws, ['시각', '구분', '메시지'],
                     [[r['time'], r['level'], r['message']] for r in log_rows],
                     st, width=70,
                     highlight=lambda v: v[1] in ('오류', '경고'))
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 8


def export_quality_report(recipe_results: list, spec_limits: dict,
                          log_rows: list, output_path: str) -> str:
    """오류·로그 전용 워크북 (전체 Recipe 기준)."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)          # 기본 빈 시트 제거
    write_quality_sheets(wb, recipe_results, spec_limits, log_rows)
    if not wb.sheetnames:
        wb.create_sheet('Summary')
    wb.save(output_path)
    return output_path
