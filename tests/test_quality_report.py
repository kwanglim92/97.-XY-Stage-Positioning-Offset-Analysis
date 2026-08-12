"""
데이터 품질 리포트 / 이상값 분류 테스트.

화면 필터(Raw Data 체크박스)와 엑셀 리포트가 **같은 기준**(core.statistics.classify_row)을
쓰는 것이 핵심이다. 두 곳이 갈라지면 "화면엔 23건인데 엑셀엔 다른 수" 같은 상황이 생기고,
계측 보고서로서 신뢰를 잃는다.
"""
import os

import pytest

from core.statistics import (classify_row, spec_bounds, axis_reference_means,
                             active_anomaly_kinds, ANOMALY_FAILED,
                             ANOMALY_OUTLIER, ANOMALY_SPEC)
from core.exporter import (collect_anomalies, build_quality_summary,
                           build_file_issues, export_quality_report)


NAN = float("nan")
SPEC = {"Vision Pattern": {"X": {"lsl": -5000.0, "usl": 5000.0},
                           "Y": {"lsl": -5000.0, "usl": 5000.0}}}


def _row(value, valid=True, outlier=False, method="X",
         lot="Lot501", site="0009_X000_Y004"):
    return {"value": value, "valid": valid, "value_valid": valid,
            "is_outlier": outlier, "method": method, "lot_name": lot,
            "site_id": site, "site_x": 0, "site_y": 4, "point_no": 1,
            "state": "COMPLETED", "filename": "Lot5_X_UL.csv"}


@pytest.fixture
def spec_on(monkeypatch):
    """Spec 초과 판정 노출 스위치를 켠다.

    Recipe별 LSL/USL의 판별 조건이 확정되지 않아 기본값은 꺼짐(SPEC_ANOMALY_ENABLED=False)
    이지만, 계산 로직 자체는 살아 있어야 하므로 테스트로 계속 고정한다.
    """
    monkeypatch.setattr("core.statistics.SPEC_ANOMALY_ENABLED", True)


def _result(rows, name="Vision Pattern", warnings=None):
    return {"short_name": name, "recipe": f"4. {name}", "round": "1st",
            "round_path": os.path.join("C:", "data", name, "1st"),
            "raw_data": rows, "data_warnings": warnings or [],
            "load_errors": []}


# ──────────────────────────────────────────────────────────────────────────
class TestClassifyRow:
    def test_normal_row_has_no_anomaly(self):
        assert classify_row(_row(1000.0), SPEC, "Vision Pattern") == []

    @pytest.mark.parametrize("row", [
        _row(NAN, valid=False),          # 장비가 기록한 측정 실패
        _row(NAN, valid=True),           # Valid=TRUE인데 값만 NaN
        _row(1000.0, valid=False),       # 값은 유한하지만 무효 표시
    ])
    def test_failed_measurement(self, row):
        assert classify_row(row, SPEC, "Vision Pattern") == [ANOMALY_FAILED]

    def test_failed_row_is_not_also_outlier_or_spec(self):
        """값이 없는 행에 이상치/Spec 판정을 붙이면 이중 집계가 된다."""
        row = _row(NAN, valid=False, outlier=True)
        assert classify_row(row, SPEC, "Vision Pattern") == [ANOMALY_FAILED]

    def test_outlier(self):
        assert classify_row(_row(1000.0, outlier=True), SPEC,
                            "Vision Pattern") == [ANOMALY_OUTLIER]

    def test_spec_violation_both_directions(self, spec_on):
        assert classify_row(_row(5000.1), SPEC, "Vision Pattern") == [ANOMALY_SPEC]
        assert classify_row(_row(-5000.1), SPEC, "Vision Pattern") == [ANOMALY_SPEC]

    def test_spec_boundary_is_inclusive(self, spec_on):
        assert classify_row(_row(5000.0), SPEC, "Vision Pattern") == []

    def test_spec_uses_deviation_when_axis_means_given(self, spec_on):
        """레시피 기준점 차이(bias)는 Stage 성능이 아니므로 판정에서 빠져야 한다.

        현장 Global Align X는 bias -3969 nm가 ±5000의 79%를 먹어, 산포가 정상인데도
        raw 기준으로는 26건이 Spec 초과로 잡혔다(편차 기준으로는 0건).
        """
        means = {"X": -3969.0}
        # 전체가 -3969 nm 밀려 있을 뿐 산포는 작은 행
        row = _row(-8000.0)
        assert classify_row(row, SPEC, "Vision Pattern") == [ANOMALY_SPEC]
        assert classify_row(row, SPEC, "Vision Pattern", axis_means=means) == []

        # bias를 걷어내도 남는 진짜 이탈은 그대로 잡힌다
        far = _row(2000.0)
        assert classify_row(far, SPEC, "Vision Pattern", axis_means=means) == [ANOMALY_SPEC]

    def test_axis_reference_means_matches_deviation_definition(self):
        """편차의 0점은 축별 '유효·유한 값 전체 평균' — PASS/FAIL과 같은 정의여야 한다."""
        data = [_row(100.0), _row(300.0), _row(NAN, valid=False),
                _row(1000.0, method="Y")]
        means = axis_reference_means(data)
        assert means["X"] == pytest.approx(200.0)   # NaN/무효 행은 제외
        assert means["Y"] == pytest.approx(1000.0)

    def test_outlier_and_spec_together(self, spec_on):
        kinds = classify_row(_row(99999.0, outlier=True), SPEC, "Vision Pattern")
        assert kinds == [ANOMALY_OUTLIER, ANOMALY_SPEC]

    def test_missing_spec_config_skips_spec_check(self, spec_on):
        """Spec 설정이 없는 Recipe는 Spec 초과로 잡지 않는다."""
        assert classify_row(_row(99999.0), SPEC, "Unknown Recipe") == []
        assert spec_bounds(SPEC, "Unknown Recipe", "X") == (None, None)

    def test_spec_is_hidden_by_default(self):
        """기본값은 꺼짐 — Recipe별 판별 조건이 정해지기 전까지 판정에 쓰지 않는다."""
        assert classify_row(_row(99999.0), SPEC, "Vision Pattern") == []
        assert ANOMALY_SPEC not in active_anomaly_kinds()
        # 이상치 판정은 스위치와 무관하게 계속 동작해야 한다
        assert classify_row(_row(99999.0, outlier=True), SPEC,
                            "Vision Pattern") == [ANOMALY_OUTLIER]

    def test_spec_logic_survives_behind_the_switch(self, spec_on):
        """스위치만 켜면 계산 로직이 그대로 되살아나야 한다 (삭제가 아니라 숨김)."""
        assert classify_row(_row(99999.0), SPEC, "Vision Pattern") == [ANOMALY_SPEC]
        assert ANOMALY_SPEC in active_anomaly_kinds()

    def test_spec_bounds_is_axis_case_insensitive(self):
        assert spec_bounds(SPEC, "Vision Pattern", "x") == (-5000.0, 5000.0)


# ──────────────────────────────────────────────────────────────────────────
class TestCollectAnomalies:
    def test_only_anomalous_rows_with_traceability(self, spec_on):
        # 정상 20건(평균 ≈ 0) + 측정 실패 1 + 편차 기준으로도 이탈하는 1건
        rows = [_row(float(v)) for v in range(-10, 10)]
        rows += [_row(NAN, valid=False), _row(9000.0)]

        found = collect_anomalies([_result(rows)], SPEC)

        assert len(found) == 2
        assert {f["kinds"] for f in found} == {ANOMALY_FAILED, ANOMALY_SPEC}
        first = next(f for f in found if f["kinds"] == ANOMALY_FAILED)
        assert first["recipe"] == "Vision Pattern"
        assert first["lot"] == "Lot501"
        # 장비 담당자가 원본을 바로 찾아갈 수 있어야 한다
        assert first["csv_path"].endswith("Lot5_X_UL.csv")
        assert "Lot501" in first["lot_dir"]

    def test_summary_counts_match_detail(self, spec_on):
        rows = [_row(float(v)) for v in range(-10, 10)]          # 정상 20
        rows += [_row(NAN, valid=False)] * 2                      # 실패 2
        rows += [_row(9000.0, outlier=True)]                      # 이상치 + Spec
        results = [_result(rows, warnings=["Lot501: MEAN U+FFFD"])]

        anomalies = collect_anomalies(results, SPEC)
        summary = build_quality_summary(results, anomalies)[0]

        assert summary["total"] == 23
        assert summary["failed"] == 2
        assert summary["outlier"] == 1
        assert summary["spec"] == 1
        assert summary["file_issues"] == 1

    def test_file_issues_include_load_errors(self):
        r = _result([], warnings=["Lot501: MEAN U+FFFD"])
        r["load_errors"] = ["Lot502 로드 실패"]
        issues = build_file_issues([r])
        assert len(issues) == 2
        assert any("로드 실패" in i["message"] for i in issues)


# ──────────────────────────────────────────────────────────────────────────
class TestExcelOutput:
    def test_workbook_has_expected_sheets_and_rows(self, tmp_path, spec_on):
        pytest.importorskip("openpyxl")
        from openpyxl import load_workbook

        rows = [_row(float(v)) for v in range(-10, 10)]
        rows += [_row(NAN, valid=False), _row(9000.0)]
        results = [_result(rows, warnings=["Lot501: MEAN U+FFFD"])]
        log_rows = [{"time": "18:56:54", "level": "경고", "message": "테스트"}]
        out = tmp_path / "QualityReport.xlsx"

        export_quality_report(results, SPEC, log_rows, str(out))

        wb = load_workbook(out)
        assert wb.sheetnames == ["Summary", "Anomalies", "By Lot", "By Die",
                                 "File Issues", "System Log"]
        assert wb["Anomalies"].max_row == 3          # 헤더 + 이상 2건
        assert wb["System Log"].max_row == 2
        assert wb["Anomalies"].freeze_panes == "A2"

    def test_empty_results_still_produces_file(self, tmp_path):
        pytest.importorskip("openpyxl")
        out = tmp_path / "Empty.xlsx"
        export_quality_report([], {}, [], str(out))
        assert out.exists()


# ──────────────────────────────────────────────────────────────────────────
class TestSystemLoggerRecords:
    def test_export_rows_keeps_time_level_message(self):
        pytest.importorskip("PySide6")
        QtWidgets = pytest.importorskip("PySide6.QtWidgets")
        os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
        if QtWidgets.QApplication.instance() is None:
            try:
                QtWidgets.QApplication([])
            except Exception:  # pragma: no cover
                pytest.skip("Qt 플랫폼 플러그인을 사용할 수 없음")

        from ui.widgets.system_logger import SystemLogger

        log = SystemLogger(QtWidgets.QTextEdit())
        log.info("정보 메시지")
        log.warn("경고 메시지")
        log.error("오류 메시지")
        log.section("구분선")

        rows = log.export_rows()
        assert [r["level"] for r in rows] == ["정보", "경고", "오류", "구분"]
        assert rows[1]["message"] == "경고 메시지"
        assert all(len(r["time"]) == 8 for r in rows)   # HH:MM:SS
